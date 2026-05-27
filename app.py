from flask import Flask, render_template, request, jsonify, send_file
import requests
import sqlite3
import json
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# 🛠️ 1. Import Groq instead of Gemini
from groq import Groq
from dotenv import load_dotenv

load_dotenv()

app = Flask(__name__)

GITHUB_TOKEN = os.getenv("GITHUB_TOKEN", "")
GROQ_API_KEY = os.getenv("GROQ_API_KEY", "")

# 🛠️ 2. Initialize the Groq Client
client = Groq(api_key=GROQ_API_KEY)

DB_PATH = "history.db"

def init_db():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS searches (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL,
            repo_count INTEGER,
            top_languages TEXT,
            generated_email TEXT,
            searched_at TEXT
        )
    """)
    conn.commit()
    conn.close()

def get_github_data(username):
    headers = {"Authorization": f"token {GITHUB_TOKEN}"} if GITHUB_TOKEN else {}
    user_resp = requests.get(f"https://api.github.com/users/{username}", headers=headers)
    if user_resp.status_code != 200:
        return None, "GitHub user not found."

    user = user_resp.json()
    repos_resp = requests.get(
        f"https://api.github.com/users/{username}/repos?per_page=50&sort=updated", 
        headers=headers
    )
    repos = repos_resp.json() if repos_resp.status_code == 200 else []

    languages = {}
    top_repo = None
    max_stars = -1

    for r in repos:
        lang = r.get("language")
        if lang:
            languages[lang] = languages.get(lang, 0) + 1
        
        stars = r.get("stargazers_count", 0)
        if stars > max_stars:
            max_stars = stars
            top_repo = {
                "name": r.get("name"),
                "description": r.get("description") or "No description provided.",
                "stars": stars,
                "url": r.get("html_url")
            }

    sorted_langs = sorted(languages.items(), key=lambda x: x[1], reverse=True)
    total_lang_count = sum(languages.values()) or 1

    profile_summary = {
        "username": username,
        "name": user.get("name") or username,
        "avatar_url": user.get("avatar_url"),
        "bio": user.get("bio"),
        "followers": user.get("followers", 0),
        "public_repos": user.get("public_repos", 0),
        "top_languages": [{"name": l[0], "percentage": round((l[1]/total_lang_count)*100)} for l in sorted_langs[:5]],
        "top_repo": top_repo
    }
    return profile_summary, None

# 🛠️ 3. Updated function to handle Groq text completion pipelines
def generate_email_groq(profile, purpose):
    if not GROQ_API_KEY:
        return (
            f"Subject: Collaboration Opportunity regarding your projects\n\n"
            f"Hi {profile['name']},\n\n"
            f"I came across your GitHub profile and noticed your excellent work. "
            f"Your focus on projects like '{profile['top_repo']['name'] if profile['top_repo'] else 'open-source software'}' looks impressive.\n\n"
            f"Best regards,\n[Your Name]"
        )

    langs_str = ", ".join([l['name'] for l in profile['top_languages']])
    repo_context = f"Their star project is '{profile['top_repo']['name']}' ({profile['top_repo']['description']})." if profile['top_repo'] else ""
    
    prompt = f"""
    You are an expert technical recruiter and open-source collaborator. Write a highly personalized, compelling, and professional cold outreach email to a developer with the following GitHub profile data:
    - Name: {profile['name']}
    - Bio: {profile['bio']}
    - Main Languages Used: {langs_str}
    - {repo_context}
    
    The outreach goal/purpose is: {purpose}.
    Keep the tone concise, authentic, developer-friendly, and free from corporate cliché. Do not include template placeholders like '[Your Name]' inside the final text. Max length: 150 words.
    """

    try:
        # Standard chat completion structure used by Groq
        completion = client.chat.completions.create(
            model="llama-3.3-70b-versatile",
            messages=[
                {"role": "user", "content": prompt}
            ],
            temperature=0.7,
            max_tokens=250
        )
        return completion.choices[0].message.content.strip()
    except Exception as e:
        print(f"!!! GROQ API ERROR DETECTED: {e}")
        return f"AI Generation skipped. Content context: Developer specializes in {langs_str}."

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/analyze", methods=["POST"])
def analyze():
    data = request.get_json() or {}
    username = data.get("username", "").strip()
    purpose = data.get("purpose", "Collaboration pitch")

    if not username:
        return jsonify({"error": "Username context cannot be blank."}), 400

    profile, error = get_github_data(username)
    if error:
        return jsonify({"error": error}), 404

    # 🛠️ 4. Direct mapping to our new function
    email = generate_email_groq(profile, purpose)
    
    langs_array = [l['name'] for l in profile['top_languages']]
    save_search(username, profile['public_repos'], langs_array, email)

    return jsonify({"profile": profile, "email": email})

def save_search(username, repo_count, top_languages, email):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute(
        "INSERT INTO searches (username, repo_count, top_languages, generated_email, searched_at) VALUES (?, ?, ?, ?, ?)",
        (username, repo_count, json.dumps(top_languages), email, datetime.now().isoformat())
    )
    conn.commit()
    conn.close()

@app.route("/history")
def history():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT username, repo_count, top_languages, searched_at FROM searches ORDER BY id DESC LIMIT 10")
    rows = c.fetchall()
    conn.close()
    return jsonify([
        {"username": r[0], "repo_count": r[1], "top_languages": json.loads(r[2]), "searched_at": r[3]} for r in rows
    ])

@app.route("/export/<username>")
def export_excel(username):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT username, repo_count, top_languages, generated_email, searched_at FROM searches WHERE username=? ORDER BY searched_at DESC LIMIT 1", (username,))
    row = c.fetchone()
    conn.close()

    if not row: return "No details tracked.", 404

    wb = Workbook()
    ws = wb.active
    ws.title = "Analytics"

    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = ["GitHub Username", "Total Public Repos", "Identified Languages Stack", "AI Strategic Email Draft", "Generated At Date"]
    for col_idx, text in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=text)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    langs_list = ", ".join(json.loads(row[2]))
    ws.append([row[0], row[1], langs_list, row[3], row[4]])

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 55
    ws.column_dimensions["E"].width = 26

    temp_filename = f"{username}_analysis.xlsx"
    wb.save(temp_filename)
    return send_file(temp_filename, as_attachment=True)

if __name__ == "__main__":
    init_db()
    app.run(debug=True, port=5000)